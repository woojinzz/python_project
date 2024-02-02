# import  openpyxl  as  op  #openpyxl 모듈 import

# wb = op.Workbook() #새로운 Workbook 객체 생성

# print(wb) #객체를 출력해보기


import  openpyxl  as  op  #openpyxl 모듈 import

# wb = op.Workbook() #새로운 워크북 객체 생성

# ws = wb.create_sheet("new_sheet1") #wb 객체를 통해 새로운 시트 생성(시트명 : new_sheet1)


# print(ws) #ws 객체 출력해보기

# wb.save("test_result.xlsx") #해당 워크북(엑셀파일) 저장하기

wb = op.load_workbook(r"C:\Users\user\Desktop\OneMinPython\test_result.xlsx")
# ws_list = wb.sheetnames  #해당 Workbook의 시트 목록을 리스트로 저장
# print(ws_list) #리스트 출력
ws = wb.active #활성화되어있는 시트 설정
print(ws)